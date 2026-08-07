#!/usr/bin/env python3
"""Prepare a trusted external component catalog artifact from an XLSX workbook.

The source workbook is read in streaming mode. Raw business fields are preserved as
strings; normalized keys and numeric values are derived into separate fields. Output
is written atomically as deterministic gzip NDJSON plus a manifest containing source,
artifact and canonical-data SHA-256 hashes.
"""

from __future__ import annotations

import argparse
import gzip
import hashlib
import io
import json
import os
import re
import tempfile
import unicodedata
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

EXPECTED_HEADERS = [
    "采集序号",
    "料号",
    "最大数量档含税单价",
    "最大数量档数量门槛",
    "商品名称",
    "品牌",
    "分类",
    "封装",
    "主要参数",
]
NUMERIC_RE = re.compile(r"^[+]?(?:\d+(?:\.\d+)?|\.\d+)$")
FORMAT_VERSION = 1


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def raw_text(value: Any) -> str | None:
    if value is None:
        return None
    return str(value)


def part_number_key(value: str) -> str:
    return value.strip().upper()


def compact_part_number_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).strip().upper()
    return "".join(normalized.split())


def parse_reference_price(raw: str) -> str | None:
    text = raw.strip().replace(",", "")
    if not NUMERIC_RE.fullmatch(text):
        raise ValueError(f"invalid non-negative decimal: {raw!r}")
    try:
        value = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"invalid decimal: {raw!r}") from exc
    if not value.is_finite() or value < 0:
        raise ValueError(f"price must be non-negative: {raw!r}")
    # 零价原样保留在 priceRaw，但不作为可报价数据；在线查询将走百炼兜底。
    return None if value == 0 else format(value, "f")


def parse_nonnegative_integer(raw: str) -> str:
    text = raw.strip().replace(",", "")
    if not NUMERIC_RE.fullmatch(text):
        raise ValueError(f"invalid quantity: {raw!r}")
    try:
        value = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"invalid quantity: {raw!r}") from exc
    if not value.is_finite() or value < 0 or value != value.to_integral_value():
        raise ValueError(f"quantity must be a non-negative integer: {raw!r}")
    return format(value.quantize(Decimal(1)), "f")


def validate_lengths(record: dict[str, Any], excel_row_no: int) -> None:
    limits = {
        "sourceSequenceRaw": 64,
        "partNumberRaw": 128,
        "partNumberKey": 128,
        "partNumberCompactKey": 128,
        "priceRaw": 32,
        "quantityThresholdRaw": 32,
        "productNameRaw": 256,
        "brandRaw": 128,
        "categoryRaw": 64,
        "packageRaw": 64,
    }
    for field, limit in limits.items():
        value = record.get(field)
        if value is not None and len(value) > limit:
            raise ValueError(
                f"row {excel_row_no}: {field} length {len(value)} exceeds {limit}"
            )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--base-name", default="external-catalog")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source = args.xlsx.resolve()
    output_dir = args.output_dir.resolve()
    if not source.is_file():
        raise SystemExit(f"source file not found: {source}")
    output_dir.mkdir(parents=True, exist_ok=True)

    artifact_path = output_dir / f"{args.base_name}.ndjson.gz"
    manifest_path = output_dir / f"{args.base_name}.manifest.json"
    source_hash = sha256_file(source)

    workbook = load_workbook(source, read_only=True, data_only=True)
    if len(workbook.worksheets) != 1:
        raise SystemExit(f"expected one worksheet, got {len(workbook.worksheets)}")
    worksheet = workbook.worksheets[0]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [raw_text(value) for value in next(rows)]
    except StopIteration as exc:
        raise SystemExit("workbook is empty") from exc
    if headers != EXPECTED_HEADERS:
        raise SystemExit(f"unexpected headers: {headers!r}")

    unique_part_keys: set[str] = set()
    canonical_hash = hashlib.sha256()
    row_count = 0
    valid_price_rows = 0

    fd, artifact_tmp_name = tempfile.mkstemp(
        dir=output_dir, prefix=f".{args.base_name}.", suffix=".ndjson.gz.tmp"
    )
    os.close(fd)
    artifact_tmp = Path(artifact_tmp_name)
    manifest_tmp = output_dir / f".{args.base_name}.manifest.json.tmp"

    try:
        with artifact_tmp.open("wb") as raw_output:
            with gzip.GzipFile(fileobj=raw_output, mode="wb", filename="", mtime=0) as gz:
                with io.TextIOWrapper(gz, encoding="utf-8", newline="\n") as text_output:
                    for excel_row_no, source_row in enumerate(rows, start=2):
                        values = list(source_row[: len(EXPECTED_HEADERS)])
                        if len(values) < len(EXPECTED_HEADERS):
                            values.extend([None] * (len(EXPECTED_HEADERS) - len(values)))
                        texts = [raw_text(value) for value in values]
                        (
                            sequence,
                            part_number,
                            price,
                            quantity,
                            product_name,
                            brand,
                            category,
                            package,
                            parameters,
                        ) = texts

                        if not part_number or not part_number.strip():
                            raise ValueError(f"row {excel_row_no}: part number is blank")
                        if price is None:
                            raise ValueError(f"row {excel_row_no}: price is blank")
                        if quantity is None:
                            raise ValueError(f"row {excel_row_no}: quantity threshold is blank")

                        key = part_number_key(part_number)
                        compact_key = compact_part_number_key(part_number)
                        price_value = parse_reference_price(price)
                        quantity_value = parse_nonnegative_integer(quantity)
                        record = {
                            "rowNo": excel_row_no,
                            "sourceSequenceRaw": sequence,
                            "partNumberRaw": part_number,
                            "partNumberKey": key,
                            "partNumberCompactKey": compact_key,
                            "priceRaw": price,
                            "priceValue": price_value,
                            "quantityThresholdRaw": quantity,
                            "quantityThresholdValue": quantity_value,
                            "productNameRaw": product_name,
                            "brandRaw": brand,
                            "categoryRaw": category,
                            "packageRaw": package,
                            "parametersRaw": parameters,
                        }
                        validate_lengths(record, excel_row_no)
                        encoded = (
                            json.dumps(record, ensure_ascii=False, separators=(",", ":"))
                            + "\n"
                        ).encode("utf-8")
                        canonical_hash.update(encoded)
                        text_output.write(encoded.decode("utf-8"))
                        row_count += 1
                        if price_value is not None:
                            valid_price_rows += 1
                        unique_part_keys.add(key)
        workbook.close()

        artifact_hash = sha256_file(artifact_tmp)
        manifest = {
            "formatVersion": FORMAT_VERSION,
            "sourceFileName": source.name,
            "sourceSha256": source_hash,
            "artifactFileName": artifact_path.name,
            "artifactSha256": artifact_hash,
            "dataSha256": canonical_hash.hexdigest(),
            "rowCount": row_count,
            "validPriceRows": valid_price_rows,
            "uniquePartKeys": len(unique_part_keys),
            "headers": EXPECTED_HEADERS,
        }
        manifest_tmp.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        os.replace(artifact_tmp, artifact_path)
        os.replace(manifest_tmp, manifest_path)
        print(
            json.dumps(
                {
                    "artifact": str(artifact_path),
                    "manifest": str(manifest_path),
                    "rowCount": row_count,
                    "uniquePartKeys": len(unique_part_keys),
                    "sourceSha256": source_hash,
                    "dataSha256": canonical_hash.hexdigest(),
                },
                ensure_ascii=False,
            )
        )
    except BaseException:
        workbook.close()
        artifact_tmp.unlink(missing_ok=True)
        manifest_tmp.unlink(missing_ok=True)
        raise


if __name__ == "__main__":
    main()
